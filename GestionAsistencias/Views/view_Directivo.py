from datetime import datetime
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from functools import wraps
from ..forms import  IngresarNuevoProfesor, IngresarNuevoHorario, PeriodoForm, PeriodoEditar, EdicionHorario, IngresarnuevoPDF, FormEditdePDF
from ..models import Directivos, Profesor, DiaAsistencia, Horario, PeriodoEscolar, Justificacion, Administrador, PDFhorario
from django.contrib import messages
import os   
from django.conf import settings
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

#funcion que sirve para verificar si un usuario es directivo
def user_is_directivo(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        # Obtener el rol del usuario desde la sesión
        rol = request.session.get('user_rol')
        
        # Verificar si el rol es 'Profesor'
        if rol == "Directivo":
            return view_func(request, *args, **kwargs)
        else:
            return render(request,"No_acceso.html")
    
    return _wrapped_view

#Funcion que sirve para entrar al inicio de un directivo
@user_is_directivo
def CasaDirectivo(request):
    id = request.session.get('user_id')
    directivo =  Directivos.objects.get(idDirectivos=id)
    return render(request, "Directivo/InicioDirectivo.html", { 'directivo' : directivo } )

#Funcion para generar el reporte
@user_is_directivo
def Reporte(request):
    busqueda = request.GET.get('search', '') #Se obtiene el parametro que en este caso es el Periodo
    periodos = PeriodoEscolar.objects.all() #obtenemos todos los periodos registrados en la base de datos
    id = request.session.get('user_id') #obtenemos el id del usuario
    directivo =  Directivos.objects.get(idDirectivos=id) #obtenemos el directivo
    profesores = Profesor.objects.filter(idDirectivos = directivo.idDirectivos) #obtenemos los profesores del directivo
    
    if busqueda: #si se ha ingresado una busqueda
        try: #se intenta buscar el periodo
            Periodo = PeriodoEscolar.objects.get(Nombre = busqueda) #se obtiene el periodo
            horarios = Horario.objects.filter(idPeriodo = Periodo, idProfesor__in = profesores) #se obtienen los horarios del periodo
            asistencias = DiaAsistencia.objects.filter(idHorario__in = horarios ) #se obtienen las asistencias del horario
        except PeriodoEscolar.DoesNotExist: #si no se encuentra el periodo
            messages.error(request, 'No se encontraron resultados') #se muestra un mensaje de error
            asistencias = DiaAsistencia.objects.none() #se obtienen las asistencias vacias
    else:
        horarios = Horario.objects.filter(idProfesor__in = profesores) #se obtienen los horarios de los profesores
        asistencias = DiaAsistencia.objects.filter(idHorario__in = horarios ) #se obtienen las asistencias del horario
    
    for profesor in profesores: #se recorre los profesores
        horarios_pro = horarios.filter(idProfesor=profesor) #se obtienen los horarios del profesor
        asistencias_pro = asistencias.filter(idHorario__in=horarios_pro) #se obtienen las asistencias del profesor
        profesor.asistencias_count = asistencias_pro.filter(Tipo="Asistencia").count() #se cuenta el numero de asistencias
        profesor.retardos_count = asistencias_pro.filter(Tipo="Retardo").count() #se cuenta el numero de retardos
    
    # Renderizar la plantilla con los datos de profesores
    return render(request, 'Directivo/ReporteDirectivo.html', {'profesores': profesores, 'periodos' : periodos })


@user_is_directivo
def GestionProfesores (request): #funcion para gestionar profesores
    id = request.session.get('user_id') #obtenemos el id del usuario
    directivo =  Directivos.objects.get(idDirectivos=id) #obtenemos el directivo
    profesores = Profesor.objects.filter(idDirectivos = directivo.idDirectivos) #obtenemos los profesores del directivo
    form = IngresarNuevoProfesor() #se crea un formulario para ingresar un nuevo profesor
    
    return render(request, 'Directivo/GestionProfesores.html', {'profesores':profesores, 'form' : form}) #se renderiza la plantilla con los profesores y el formulario



def buscar_profesores(request): #funcion para buscar profesores
    query = request.GET.get('q', '')  # Obtiene el texto de búsqueda
    id = request.session.get('user_id') #obtenemos el id del usuario
    directivo =  Directivos.objects.get(idDirectivos=id) #obtenemos el directivo
    if query: #si se ha ingresado una busqueda
        profesores = Profesor.objects.filter( 
            Matricula__icontains=query, #buscamos por matricula
            idDirectivos = directivo.idDirectivos #solo buscamos profesores del directivo
        )
    else: #si no se ha ingresado una busqueda
        profesores = Profesor.objects.filter(idDirectivos = directivo.idDirectivos) #obtenemos los profesores del directivo

    data = [ #se crea un arreglo con los datos de los profesores
        { #se crea un objeto con los datos de cada profesor
            'id': profesor.idProfesor, #obtenemos el id del profesor
            'nombre': profesor.Nombre, #obtenemos el nombre del profesor
            'apellidos': profesor.Apellidos, #obtenemos los apellidos del profesor
            'matricula': profesor.Matricula, #obtenemos la matricula del profesor
            'correo': profesor.Correo, #obtenemos el correo del profesor
            'contrasena': profesor.Contrasena , #obtenemos la contraseña del profesor
        } for profesor in profesores #se recorre los profesores
    ] 
    return JsonResponse(data, safe=False) #se envia la respuesta en formato json

@user_is_directivo 
def crear_profesor(request): #funcion para crear un nuevo profesor
    if request.method == 'POST': #si se ha enviado el formulario
        form = IngresarNuevoProfesor(request.POST) #se crea un formulario con los datos del profesor
        if form.is_valid(): #si el formulario es valido
            id  = request.session.get('user_id') #obtenemos el id del usuario
            directivo = Directivos.objects.get(idDirectivos=id) #obtenemos el directivo
            profesor = form.save(commit=False) #se crea un nuevo profesor
            profesor.idDirectivos = directivo #se asigna el directivo al profesor
            profesor.save() #se guarda el profesor

            return redirect('GestionProfesores') #se redirecciona a la pagina de gestion de profesores

    return redirect('GestionProfesores') #si no se ha enviado el formulario se redirecciona a la pagina de gestion de profesores





@user_is_directivo
def BorrarProfesor (request, id): #funcion para borrar un profesor
    profesor = get_object_or_404(Profesor, idProfesor=id) #obtenemos el profesor
    profesor.delete() #se borra el profesor
    return redirect('GestionProfesores') #se redirecciona a la pagina de gestion de profesores

@user_is_directivo
def  EditarProfesor (request, id): #funcion para editar un profesor
    profesor = get_object_or_404(Profesor, idProfesor=id) #obtenemos el profesor
    
    if request.method == 'POST': #si se ha enviado el formulario
        # Obtén los datos del formulario
        profesor.Nombre = request.POST.get('nombre')
        profesor.Apellidos = request.POST.get('apellido')
        profesor.Matricula = request.POST.get('matricula')
        profesor.Correo  = request.POST.get('Correo')
        profesor.Contrasena  = request.POST.get('Contrasena')

        # Guarda los cambios
        profesor.save()
        
        # Redirige a otra página
        return redirect('GestionProfesores')

    # Si es un método GET, muestra el formulario con los datos del profesor
    return redirect('GestionProfesores')


def CerrarSesionDirectivo(request):
    # Eliminar datos específicos de la sesión
    if 'user_id' in request.session:
        del request.session['user_id']
    if 'user_rol' in request.session:   
        del request.session['user_rol']
    return redirect("Inicio")

@user_is_directivo
def GestionHorariosPDF(request, id): #funcion para generar un pdf con los horarios de un profesor
    profesor = get_object_or_404(Profesor, idProfesor=id) #obtenemos el profesor
    horarios = Horario.objects.filter(idProfesor = profesor) #obtenemos los horarios del profesor
    horariosPDF = PDFhorario.objects.filter(idHorario__in = horarios) #obtenemos los horarios pdf del profesor
    form = IngresarnuevoPDF() #se crea un formulario para ingresar un nuevo pdf
    form2 = EdicionHorario() #se crea un formulario para editar un horario
    
    return render(request, 'Directivo/GestionHorariosPDF.html', {'horarios' : horarios, 'PDFs' : horariosPDF, 'form' : form, 'profesor':profesor, 'form2' : form2}) #se muestra la pagina de gestion de horarios pdf


@user_is_directivo
def GestionHorarios(request, id): #funcion para gestionar los horarios de un profesor
    profesor = get_object_or_404(Profesor, idProfesor=id) #obtenemos el profesor
    horarios = Horario.objects.filter(idProfesor = profesor) #obtenemos los horarios del profesor
    horariosPDF = PDFhorario.objects.filter(idHorario__in = horarios) #obtenemos los horarios pdf del profesor
    
    for horario in horarios:
        horario.has_pdf = any(pdf.idHorario.idHorario == horario.idHorario for pdf in horariosPDF) #se verifica si el horario tiene un pdf asociado
    
    form = IngresarNuevoHorario() #se crea un formulario para ingresar un nuevo horario
    form2 = EdicionHorario() #se crea un formulario para editar un horario
    form3 = IngresarnuevoPDF() #se crea un formulario para ingresar un nuevo pdf
    form4 = FormEditdePDF() #se crea un formulario para editar un pdf
    
    return render(request, 'Directivo/GestionHorarios.html', {'horarios' : horarios, 'horariosPDF': horariosPDF, 'form' : form, 'profesor':profesor, 'form2' : form2, 'form3' : form3, 'form4': form4}) #se muestra la pagina de gestion de horarios



@user_is_directivo
def EliminarHorario(request,id,idPro):
    horario = get_object_or_404(Horario, idHorario=id) #obtenemos el horario
    horario.delete() #eliminamos el horario
    
    return redirect('GestionHorarios', id =  idPro) #redirigimos a la pagina de gestion de horarios

@user_is_directivo
def CrearHorario(request,id): #funcion para crear un nuevo horario
    if request.method == 'POST': #si se envia un formulario
        form = IngresarNuevoHorario(request.POST) #se crea un formulario para ingresar un nuevo horario
        if form.is_valid(): #si el formulario es valido
            profesor =  get_object_or_404(Profesor, idProfesor=id) #obtenemos el profesor
            horario = form.save(commit=False) #se crea un nuevo horario
            horario.idProfesor = profesor #se le asigna el profesor al horario
            horario.save() #se guarda el horario
            return redirect('GestionHorarios',id = id) #redirigimos a la pagina de gestion de horarios
        else: #si el formulario no es valido
            return redirect('GestionHorarios',id = id) #redirigimos a la pagina de gestion de horarios
    return redirect('GestionHorarios',id = id) #redirigimos a la pagina de gestion de horarios

@user_is_directivo
def EditarHorario(request, id):
    # Obtener el horario que se va a editar
    horario = get_object_or_404(Horario, idHorario=id)
    idPro =  horario.idProfesor.idProfesor

    if request.method == 'POST':
        # Crear una instancia del formulario con los datos del POST
        form = IngresarNuevoHorario(request.POST, instance=horario)
        
        if form.is_valid():
            # Guardar la instancia actualizada
            form.save()
            return redirect('GestionHorarios',id = idPro)  # Redirigir a la vista de gestión de horarios
    else:
        # Si no es una solicitud POST, se muestra el formulario con los datos existentes
        return redirect('GestionHorarios',id = idPro) 

    # Renderizar la plantilla con el formulario
    return redirect('GestionHorarios',id = idPro)


@user_is_directivo
def buscar_periodos(request): #funcion para buscar periodos
    query = request.GET.get('q', '')  # Obtiene el texto de búsqueda
    if query: # Si hay texto de búsqueda
        periodos = PeriodoEscolar.objects.filter(
            Nombre__icontains=query # Busca en el campo Nombre
        )
    else:
        periodos = PeriodoEscolar.objects.all() # Si no hay texto de búsqueda, devuelve todos los periodos

    data = [ # Prepara los datos para la plantilla
        { # Cada elemento es un diccionario con los datos del periodo
            'id': periodo.idPeriodo, # Id del periodo
            'nombre': periodo.Nombre, # Nombre del periodo
            'fechainicio': periodo.FechaInicio.strftime('%Y-%m-%d'), # Fecha de inicio del periodo
            'fechafin': periodo.FechaFin.strftime('%Y-%m-%d') # Fecha de fin del periodo
        }
        for periodo in periodos # Crea un elemento por cada periodo encontrado
    ]
    return JsonResponse(data, safe=False) # Devuelve los datos en formato JSON


@user_is_directivo
def GestionPeriodos(request): #funcion para gestionar periodos
    periodos = PeriodoEscolar.objects.all() # Obtiene todos los periodos
    form = PeriodoForm() # Crea un formulario vacío para el periodo
    form2 = PeriodoEditar() # Crea un formulario vacío para editar el periodo
    return render(request, 'Directivo/GestionPeriodos.html', {'periodos': periodos, 'form': form, 'form2': form2}) # Renderiza la plantilla con los periodos y los formularios

@user_is_directivo
def EliminarPeriodo (request,id):
    # Obtener el periodo que se va a eliminar
    periodo = get_object_or_404(PeriodoEscolar, idPeriodo=id) # Obtiene el periodo con el id especificado o devuelve un 404 si no existe
    periodo.delete() # Elimina el periodo
    return redirect('GestionPeriodos') # Redirige a la vista de gestión de periodos

@user_is_directivo
def CrearPeriodo (request): #funcion para crear un nuevo periodo
    if request.method == 'POST':
        # Crear una instancia del formulario con los datos del POST
        form = PeriodoForm(request.POST) # Crea un formulario con los datos del POST
        if  form.is_valid():
            # Guardar la instancia actualizada
            form.save() # Guarda el formulario
            return redirect('GestionPeriodos') # Redirige a la vista de gestión de periodos
        else: # Si el formulario no es válido
            return redirect('GestionPeriodos') # Redirige a la vista de gestión de periodos
    return redirect('GestionPeriodos') # Redirige a la vista de gestión de periodos


@user_is_directivo
def  EditarPeriodo (request,id): #funcion para editar un periodo
    # Obtener el periodo que se va a editar
    periodo = get_object_or_404(PeriodoEscolar, idPeriodo=id) # Obtiene el periodo con el id especificado o devuelve un 404 si no existe
    
    if  request.method == 'POST': # Si el método de la petición es POST
        # Crear una instancia del formulario con los datos del POST
        form = PeriodoForm(request.POST, instance=periodo) 
        
        if  form.is_valid():
            # Guardar la instancia actualizada
            form.save()
            return redirect('GestionPeriodos')
    else:
        return redirect('GestionPeriodos')

@user_is_directivo
def crear_pdf(request, id, idHorario): #funcion para crear un pdf
    if request.method == 'POST': # Si el método de la petición es POST
        form = IngresarnuevoPDF(request.POST, request.FILES) # Crea un formulario con los datos del POST y los archivos subidos
        if form.is_valid(): # Si el formulario es válido
            pdf_horario = form.save(commit=False) # Guarda el formulario sin guardar en la base de datos
            pdf_horario.FechaModificacion = timezone.now() # Establece la fecha de modificación
            horario = Horario.objects.get(idHorario = idHorario) # Obtiene el horario con el id especificado
            pdf_horario.idHorario = horario # Establece el horario
            pdf_horario.save() # Guarda el formulario en la base de datos
            return redirect('GestionHorarios', id=id)  # Redirige usando el `idProfesor`
        else:
            # Renderiza de nuevo la página con el formulario y los errores
            return redirect('GestionHorarios', id=id) # Redirige usando el `idProfesor`

    else:
        return redirect('GestionHorarios', id=id) # Redirige usando el `idProfesor`
    

    
@user_is_directivo    
def EliminarPDF(reques, id, idPro): #funcion para eliminar un pdf
    PDF = get_object_or_404(PDFhorario, idPDFhorario=id)  # Obtiene el pdf con el id especificado o devuelve un 404 si no existe
    
    if PDF.horario_pdf: # Si el pdf tiene un horario asociado
        file_path = os.path.join(settings.MEDIA_ROOT, PDF.horario_pdf.name) # Obtiene el path del archivo
        if os.path.isfile(file_path): # Si el archivo existe
            os.remove(file_path) # Elimina el archivo
            
            
    PDF.delete() # Elimina el pdf de la base de datos
    return redirect('GestionHorarios', id=idPro) # Redirige usando el `idProfesor`

    
@user_is_directivo
def EditarPDF(request, id, idPro): #funcion para editar un pdf
    pdf = get_object_or_404(PDFhorario, idPDFhorario=id) # Obtiene el pdf con el id especificado o devuelve un 404 si no existe
    if request.method == 'POST': # Si el método de la petición es POST
        form = IngresarnuevoPDF(request.POST, request.FILES, instance=pdf) # Crea un formulario con los datos del POST y los archivos subidos
        if form.is_valid(): # Si el formulario es válido
            # Actualiza la fecha de modificación
            pdf.FechaModificacion = timezone.now()
            pdf = form.save(commit=False) # Guarda el formulario sin guardar en la base de datos
            pdf.save()   # Guarda el formulario en la base de datos
            return redirect('GestionHorarios', id=idPro) # Redirige usando el `idProfesor`
    return redirect('GestionHorarios', id=idPro) # Redirige usando el `idProfesor`



@user_is_directivo
def Justificante (request): #funcion para ver los justificantes
    
    id = request.session.get('user_id') # Obtiene el id del usuario
    
    directivo =  Directivos.objects.get(idDirectivos=id) # Obtiene el directivo con el id especificado
    profesores = Profesor.objects.filter(idDirectivos = directivo.idDirectivos) # Obtiene los profesores del directivo
    horarios =  Horario.objects.filter(idProfesor__in = profesores) # Obtiene los horarios de los profesores

    asistencias = DiaAsistencia.objects.filter(idHorario__in = horarios) # Obtiene las asistencias de los horarios
    justificantes  = Justificacion.objects.filter(idDiaAsistencia__in=asistencias) # Obtiene los justificantes de las asistencias
    
    for justificante in justificantes: # Recorre los justificantes
        profesor = Profesor.objects.get(idProfesor=justificante.idDiaAsistencia.idHorario.idProfesor.idProfesor) # Obtiene el profesor del justificante
        justificante.profesor = profesor.Matricula # Asigna el matricula del profesor al justificante
    
    return render (request, 'Directivo/Justificantes.html',{'justificantes':justificantes}) # Renderiza la vista con los justificantes


@user_is_directivo    
def AceptarJustificante (request, id): #funcion para aceptar un justificante
    
    justificante = get_object_or_404(Justificacion, idJustificacion=id) # Obtiene el justificante con el id especificado o devuelve un 404 si no
    justificante.estado = 'Aprobado' # Cambia el estado del justificante a 'Aprobado'
    justificante.save() # Guarda el cambio en la base de datos
    asistencia = get_object_or_404(DiaAsistencia, id =  justificante.idDiaAsistencia.id) # Obtiene la asistencia del justificante
    asistencia.Tipo = 'Asistencia'; # Cambia el tipo de la asistencia a 'Asistencia'
    asistencia.save() # Guarda el cambio en la base de datos
    
    return redirect ('Justificantes') # Redirige a la vista de justificantes
    

    
@user_is_directivo
def validar_nombre(request):
    nombre = request.GET.get('nombre')
    directivo_id = request.GET.get('id')  # ID del directivo que estamos editando
    print ("hola {directivo_id}")
    
    
    if nombre:
        # Excluir el `directivo_id` solo en `Directivos`
        if directivo_id:
            existe = (
                PeriodoEscolar.objects.filter(Nombre=nombre).exclude(idPeriodo=directivo_id).exists()
            )
        else:
            existe = (
                PeriodoEscolar.objects.filter(Nombre=nombre).exists()
            )
        return JsonResponse({'existe': existe})
    
    # Respuesta si el `correo` no se proporciona
    return JsonResponse({'error': 'Correo no proporcionado'}, status=400)


@user_is_directivo
def validar_fechas(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_final = request.GET.get('fecha_final')
    
   
    # Validar que ambas fechas están presentes
    if fecha_inicio and fecha_final:
        try:
            # Convertir las fechas a objetos datetime
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_final_dt = datetime.strptime(fecha_final, '%Y-%m-%d')

            # Comparar las fechas
            es_valido = fecha_inicio_dt <= fecha_final_dt
            return JsonResponse({'es_valido': es_valido})
        except ValueError:
            # Si hay un error de formato, retorna que no es válido
            return JsonResponse({'es_valido': False})
    else:
        # Retorna que no es válido si alguna fecha falta
        return JsonResponse({'es_valido': False})

@user_is_directivo    
def validar_matricula_profesor(request):
    matricula = request.GET.get('matricula')
    profesor_id = request.GET.get('id')  # ID del directivo que estamos editando
    if matricula:
        # Excluir el `directivo_id` solo en `Directivos`
        if profesor_id:
            existe = (
                Directivos.objects.filter(Matricula=matricula).exists() or
                Administrador.objects.filter(Matricula=matricula).exists() or
                Profesor.objects.filter(Matricula=matricula).exclude(idProfesor=profesor_id).exists()
            )
        else:
            existe = (
                Directivos.objects.filter(Matricula=matricula).exists() or
                Administrador.objects.filter(Matricula=matricula).exists() or
                Profesor.objects.filter(Matricula=matricula).exists()
            )
        return JsonResponse({'existe': existe})
    
    
    


def validar_correo_profesor(request): 
    correo = request.GET.get('correo')
    profesor_id = request.GET.get('id')  # ID del directivo que estamos editando
    print (profesor_id)
    if correo:
        # Excluir el `directivo_id` solo en `Directivos`
        if profesor_id:
            existe = (
                Directivos.objects.filter(Correo=correo).exists() or
                Profesor.objects.filter(Correo=correo).exclude(idProfesor=profesor_id).exists()
            )
        else:
            existe = (
                Directivos.objects.filter(Correo=correo).exists() or
                Profesor.objects.filter(Correo=correo).exists()
            )
        return JsonResponse({'existe': existe})
    
    # Respuesta si el `correo` no se proporciona
    return JsonResponse({'error': 'Correo no proporcionado'}, status=400)




def validar_periodo_profesor(request): # Validar si el profesor ya tiene un periodo asignado
    id_periodo = request.GET.get('idPeriodo') # ID del periodo que estamos editando
    profesor_id = request.GET.get('profesorId') # ID del profesor que estamos editando
    id_horario = request.GET.get('id') # ID del horario que estamos editando
    
    if id_horario: # Si el horario tiene un profesor asignado
        if Horario.objects.filter(idPeriodo=id_periodo, idProfesor=profesor_id).exclude(idHorario =id_horario).exists(): # Si el profesor ya tiene un periodo asignado
            return JsonResponse({'existe': True}) # Si el profesor ya tiene un periodo asignado
    else: # Si el horario no tiene un profesor asignado
        if Horario.objects.filter(idPeriodo=id_periodo, idProfesor=profesor_id).exists(): # Si el profesor ya tiene un periodo asignado
            return JsonResponse({'existe': True}) # Si el profesor ya tiene un periodo asignado
    return JsonResponse({'existe': False}) # Si el profesor no tiene un periodo asignado




def generar_reporte_pdf(request):
    busqueda = request.GET.get('search', '')
    id = request.session.get('user_id')
    directivo =  Directivos.objects.get(idDirectivos=id)
    profesores = Profesor.objects.filter(idDirectivos = directivo.idDirectivos)
    
    if busqueda:
        try:
            Periodo = PeriodoEscolar.objects.get(Nombre = busqueda)
            horarios = Horario.objects.filter(idPeriodo = Periodo, idProfesor__in = profesores)
            asistencias = DiaAsistencia.objects.filter(idHorario__in = horarios )    
        except PeriodoEscolar.DoesNotExist:
            messages.error(request, 'No se encontraron resultados')
            asistencias = DiaAsistencia.objects.none()
    else:
        horarios = Horario.objects.filter(idProfesor__in = profesores)
        asistencias = DiaAsistencia.objects.filter(idHorario__in = horarios )
    
    for profesor in profesores:
        horarios_pro = horarios.filter(idProfesor=profesor)
        asistencias_pro = asistencias.filter(idHorario__in=horarios_pro)
        profesor.asistencias_count = asistencias_pro.filter(Tipo="Asistencia").count()
        profesor.retardos_count = asistencias_pro.filter(Tipo="Retardo").count()

    # Crear el PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_asistencias.pdf"'
    
    pdf = canvas.Canvas(response, pagesize=letter)
    pdf.setFont("Helvetica", 12)

    # Título del reporte
    pdf.drawString(200, 750, "Reporte de Asistencias")
    pdf.drawString(100, 730, f"Periodo: {busqueda if busqueda else 'Todos los periodos'}")
    
    # Encabezados de la tabla
    pdf.drawString(50, 700, "Matrícula")
    pdf.drawString(200, 700, "Asistencias")
    pdf.drawString(350, 700, "Retardos")

    # Datos de los profesores
    y = 680  # Altura inicial
    for profesor in profesores:
        pdf.drawString(50, y, str(profesor.Matricula))
        pdf.drawString(200, y, str(profesor.asistencias_count))
        pdf.drawString(350, y, str(profesor.retardos_count))
        y -= 20
        if y < 50:  # Salto de página si hay demasiados datos
            pdf.showPage()
            pdf.setFont("Helvetica", 12)
            y = 750

    pdf.save()
    return response



from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.contrib import messages

def generar_reporte_excel(request):
    # Obtener los datos necesarios
    busqueda = request.GET.get('search', '')
    id = request.session.get('user_id')
    directivo = Directivos.objects.get(idDirectivos=id)
    profesores = Profesor.objects.filter(idDirectivos=directivo.idDirectivos)

    if busqueda:
        try:
            periodo = PeriodoEscolar.objects.get(Nombre=busqueda)
            horarios = Horario.objects.filter(idPeriodo=periodo, idProfesor__in=profesores)
            asistencias = DiaAsistencia.objects.filter(idHorario__in=horarios)
        except PeriodoEscolar.DoesNotExist:
            messages.error(request, 'No se encontraron resultados')
            asistencias = DiaAsistencia.objects.none()
    else:
        horarios = Horario.objects.filter(idProfesor__in=profesores)
        asistencias = DiaAsistencia.objects.filter(idHorario__in=horarios)

    # Contar asistencias y retardos por profesor
    for profesor in profesores:
        horarios_pro = horarios.filter(idProfesor=profesor)
        asistencias_pro = asistencias.filter(idHorario__in=horarios_pro)
        profesor.asistencias_count = asistencias_pro.filter(Tipo="Asistencia").count()
        profesor.retardos_count = asistencias_pro.filter(Tipo="Retardo").count()

    # Crear un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Asistencias"

    # Título del reporte
    ws.merge_cells('A1:D1')
    ws['A1'] = "Reporte de Asistencias"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Subtítulo con el período
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Periodo: {busqueda if busqueda else 'Todos los periodos'}"
    ws['A2'].font = Font(size=12, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Encabezados de la tabla
    headers = ["Matrícula", "Nombre", "Asistencias", "Retardos"]
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Agregar los datos de los profesores
    for profesor in profesores:
        ws.append([
            profesor.Matricula,
            profesor.Nombre,
            profesor.asistencias_count,
            profesor.retardos_count,
        ])

    # Ajustar ancho de las columnas
    column_widths = [15, 25, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Configurar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_asistencias.xlsx"'

    # Guardar el archivo en la respuesta
    wb.save(response)
    return response
